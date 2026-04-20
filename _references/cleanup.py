"""Generate cleaned Sales Tracker workbook with structured Products, Customers, Suppliers sheets."""
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '_references/Sales Tracker 2025.xlsx'
OUT = '_references/Sales Tracker 2025 - cleaned.xlsx'

# Brand prefix map (conservative — only where confident from the data)
PREFIX_BRAND = {
    'BBR': 'B.Braun',
    'EUR': 'Euromed',
    'TER': 'Terumo',
    'IND': 'Indoplas',
    'ORM': 'Ormed',
    'ORE': 'Orex',
    'SIM': 'Simplex',
    'MUL': 'Multicare',
    'SURGU': 'Sureguard',
}

CATEGORY_PREFIX = {
    'Medical': 'MED',
    'Utility': 'UTL',
    'Office': 'OFC',
    'Spa_Salon': 'SPA',
    'Kitchen': 'KIT',
}

def parse_bracket(name: str):
    """Extract content inside [...] from item name. Returns (clean_name, variant_raw)."""
    if not isinstance(name, str):
        return name, None
    m = re.search(r'\[([^\]]+)\]', name)
    if not m:
        return name.strip(), None
    variant = m.group(1).strip()
    clean = re.sub(r'\s*\[[^\]]+\]\s*', ' ', name).strip()
    clean = re.sub(r'\s+', ' ', clean)
    return clean, variant

def infer_uom(name: str, variant: str | None):
    """Guess UOM from bracket/variant text or item name keywords."""
    text = f"{variant or ''} {name or ''}".lower()
    for kw, uom in [
        ('case', 'case'), ('box', 'box'), ('pack', 'pack'),
        ('bottle', 'bottle'), ('vial', 'vial'), ('roll', 'roll'),
        ('bag', 'bag'), ('tube', 'tube'), ('pair', 'pair'),
    ]:
        if kw in text:
            return uom
    return ''

def infer_brand(row):
    existing = row.get('Brand')
    if pd.notna(existing) and str(existing).strip():
        return str(existing).strip().title()
    name = str(row.get('Items', ''))
    m = re.match(r'^([A-Z]{2,6})\s', name)
    if m and m.group(1) in PREFIX_BRAND:
        return PREFIX_BRAND[m.group(1)]
    return ''

def clean_outlet(val):
    if pd.isna(val):
        return ''
    # Split on '/' and take first (primary) supplier, normalize
    first = str(val).split('/')[0].strip()
    return first.title() if first else ''

# ============ PRODUCTS ============
prod = pd.read_excel(SRC, sheet_name='Pricelist2025v2')
prod = prod[['Supplies', 'Items', 'Price', 'Brand', 'Outlet', 'Cost']].copy()
prod = prod.dropna(subset=['Items'])
prod['Items'] = prod['Items'].astype(str).str.strip()

parsed = prod['Items'].apply(parse_bracket)
prod['Clean Name'] = parsed.apply(lambda x: x[0])
prod['Variant/Pack'] = parsed.apply(lambda x: x[1])
prod['UOM'] = prod.apply(lambda r: infer_uom(r['Clean Name'], r['Variant/Pack']), axis=1)
prod['Brand (resolved)'] = prod.apply(infer_brand, axis=1)
prod['Preferred Supplier'] = prod['Outlet'].apply(clean_outlet)
prod['Category'] = prod['Supplies'].fillna('').astype(str).str.strip()

# Generate SKU: <CAT>-<NNNN>
prod = prod.sort_values(['Category', 'Clean Name']).reset_index(drop=True)
skus = []
counters = {}
for _, r in prod.iterrows():
    prefix = CATEGORY_PREFIX.get(r['Category'], 'GEN')
    counters[prefix] = counters.get(prefix, 0) + 1
    skus.append(f"{prefix}-{counters[prefix]:04d}")
prod['SKU'] = skus

products_out = prod[[
    'SKU', 'Clean Name', 'Brand (resolved)', 'Category',
    'Variant/Pack', 'UOM', 'Price', 'Cost',
    'Preferred Supplier', 'Items',
]].rename(columns={
    'Clean Name': 'Name',
    'Brand (resolved)': 'Brand',
    'Price': 'Sales Price',
    'Items': 'Original Name',
})
products_out['Active'] = 'Yes'
products_out['Notes'] = ''
products_out = products_out[[
    'SKU', 'Name', 'Brand', 'Category', 'Variant/Pack', 'UOM',
    'Sales Price', 'Cost', 'Preferred Supplier', 'Active',
    'Original Name', 'Notes',
]]

# ============ CUSTOMERS ============
cust = pd.read_excel(SRC, sheet_name='Client List')
cust.columns = [str(c).strip() for c in cust.columns]
cust = cust.rename(columns={
    'Unnamed: 0': 'Branch',
    'Company name': 'Company Name',
    'Contact Details': 'Mobile',
})
cust = cust.dropna(subset=['Company Name'])

def format_mobile(v):
    if pd.isna(v):
        return ''
    s = re.sub(r'\D', '', str(v))
    if s.startswith('63'):
        s = s[2:]
    if s.startswith('9') and len(s) == 10:
        return f"+63 {s[:3]} {s[3:6]} {s[6:]}"
    return str(v)

cust['Mobile'] = cust['Mobile'].apply(format_mobile)
cust['Customer Code'] = [f"CUST-{i+1:03d}" for i in range(len(cust))]
cust['Email'] = ''
cust['Payment Terms'] = 'Net 15'  # default assumption — user to adjust
cust['Account Status'] = 'Active'
cust['Notes'] = ''

customers_out = cust[[
    'Customer Code', 'Company Name', 'Branch', 'Address', 'TIN',
    'Contact Person', 'Mobile', 'Email', 'Payment Terms',
    'Account Status', 'Notes',
]]

# ============ SUPPLIERS ============
sup_raw = pd.read_excel(SRC, sheet_name='Supplier List')
sup_raw = sup_raw.dropna(subset=['Supplier'])
sup_raw = sup_raw[~sup_raw['Supplier'].astype(str).str.contains(r'^\s*New:?\s*$', na=False)]

# Merge with outlets referenced in pricelist
used_outlets = set(prod['Preferred Supplier'].dropna().unique()) - {''}
existing_names = set(sup_raw['Supplier'].astype(str).str.strip().str.title())

all_suppliers = sorted(existing_names | used_outlets)

sup_lookup = sup_raw.set_index(sup_raw['Supplier'].astype(str).str.strip().str.title())

rows = []
for i, name in enumerate(all_suppliers, 1):
    contact = ''
    if name in sup_lookup.index:
        val = sup_lookup.loc[name, 'Name']
        if pd.notna(val):
            contact = str(val)
    # Brands supplied — derive from products where this outlet is primary
    brands = prod[prod['Preferred Supplier'] == name]['Brand (resolved)'].dropna()
    brands = sorted(set(b for b in brands if b))
    sku_count = int((prod['Preferred Supplier'] == name).sum())
    rows.append({
        'Supplier Code': f"SUP-{i:03d}",
        'Supplier Name': name,
        'Contact Person': contact,
        'Mobile': '',
        'Email': '',
        'Address': '',
        'Brands Supplied': ', '.join(brands),
        'SKUs Linked': sku_count,
        'Notes': '',
    })
suppliers_out = pd.DataFrame(rows)

# ============ WRITE WORKBOOK ============
wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill('solid', start_color='305496')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name='Arial', size=10)
TO_FILL = PatternFill('solid', start_color='FFF2CC')
BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def write_sheet(name, df, highlight_cols=None, currency_cols=None, widths=None):
    ws = wb.create_sheet(name)
    highlight_cols = set(highlight_cols or [])
    currency_cols = set(currency_cols or [])

    # Header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    # Body
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, (col_name, val) in enumerate(zip(df.columns, row), 1):
            if pd.isna(val):
                val = ''
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            if col_name in currency_cols and isinstance(val, (int, float)) and val != '':
                cell.number_format = '"₱"#,##0.00'
            if col_name in highlight_cols and (val == '' or val is None):
                cell.fill = TO_FILL

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    if widths:
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

write_sheet(
    'Products', products_out,
    highlight_cols={'Sales Price', 'Cost', 'Preferred Supplier', 'Brand', 'UOM'},
    currency_cols={'Sales Price', 'Cost'},
    widths={'A': 12, 'B': 42, 'C': 14, 'D': 12, 'E': 24, 'F': 10,
            'G': 12, 'H': 12, 'I': 18, 'J': 8, 'K': 42, 'L': 24},
)
write_sheet(
    'Customers', customers_out,
    highlight_cols={'Email', 'Contact Person', 'Address', 'TIN', 'Mobile'},
    widths={'A': 12, 'B': 34, 'C': 16, 'D': 48, 'E': 20, 'F': 18,
            'G': 18, 'H': 28, 'I': 14, 'J': 14, 'K': 24},
)
write_sheet(
    'Suppliers', suppliers_out,
    highlight_cols={'Contact Person', 'Mobile', 'Email', 'Address'},
    widths={'A': 12, 'B': 22, 'C': 18, 'D': 18, 'E': 28, 'F': 40,
            'G': 30, 'H': 12, 'I': 24},
)

# README sheet
ws = wb.create_sheet('README', 0)
readme = [
    ('V&V Sales Tracker — Cleaned', ''),
    ('', ''),
    ('Sheet', 'Purpose'),
    ('Products', f'Product catalog — {len(products_out)} SKUs from Pricelist2025v2'),
    ('Customers', f'Customer accounts — {len(customers_out)} companies from Client List'),
    ('Suppliers', f'Supplier directory — {len(suppliers_out)} suppliers (merged from Supplier List + Pricelist Outlets)'),
    ('', ''),
    ('Legend', ''),
    ('Yellow-highlighted cells', 'Missing data — fill these in before Odoo import'),
    ('Original Name (Products)', 'Kept for reference; Name column is cleaned'),
    ('Variant/Pack (Products)', 'Raw bracket content from original names — refine manually where needed'),
    ('Payment Terms (Customers)', 'Defaulted to "Net 15" — change to "Prepaid" for non-loyal customers'),
    ('', ''),
    ('Next steps', ''),
    ('1', 'Fill all yellow cells (missing prices, emails, supplier contacts, etc.)'),
    ('2', 'Review Variant/Pack column — some may need manual cleanup'),
    ('3', 'Set correct Payment Terms per customer (Net 15 vs Prepaid)'),
    ('4', 'Upload to Google Sheets OR use directly for Odoo import'),
]
for r, (a, b) in enumerate(readme, 1):
    ws.cell(row=r, column=1, value=a).font = Font(name='Arial', bold=(r in [1, 3, 8, 14]), size=11 if r == 1 else 10)
    ws.cell(row=r, column=2, value=b).font = Font(name='Arial', size=10)
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 80

wb.save(OUT)

# Report
missing_price = int(products_out['Sales Price'].isna().sum() + (products_out['Sales Price'] == '').sum())
print(f"Saved: {OUT}")
print(f"Products: {len(products_out)} (missing prices: {products_out['Sales Price'].isna().sum()})")
print(f"Customers: {len(customers_out)}")
print(f"Suppliers: {len(suppliers_out)}")
print(f"Categories:")
print(products_out['Category'].value_counts().to_string())
